# src/processor.py
# Este módulo contiene la LÓGICA DE NEGOCIO pura.
# Se encarga de manipular el archivo Excel usando la librería 'openpyxl'.

from openpyxl import load_workbook  # Para cargar archivos existentes
from openpyxl.utils import get_column_letter  # Helper para convertir índice 1 -> 'A', 2 -> 'B'
import os  # Para operaciones del sistema (verificar si archivo existe)

def ejecutar_accion(instruccion, file_path):
    """
    Recibe una instrucción estructurada y la ejecuta sobre el archivo Excel.
    
    Args:
        instruccion (dict): Diccionario con la orden (action, column, mode, etc.)
        file_path (str): Ruta al archivo .xlsx
        
    Returns:
        str: Mensaje de resultado para el usuario.
    """
    
    # Verificamos que el archivo exista antes de intentar abrirlo
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"El archivo no existe: {file_path}")

    # Cargamos el libro (Workbook) y seleccionamos la hoja activa
    wb = load_workbook(file_path)
    ws = wb.active

    # --- CASO 1: LIMPIEZA DE COLUMNA ---
    if instruccion["action"] == "clean_column":
        col = instruccion["column"]
        mode = instruccion.get("mode", "keep_digits") # Modo de limpieza (por defecto keep_digits si falla IA)
        
        count = 0        # Filas recorridas
        modificadas = 0  # Celdas que realmente cambiaron
        
        # Obtenemos la última fila con datos para saber hasta dónde iterar.
        max_row = ws.max_row
        if max_row < 2:
            return "El archivo parece estar vacío o solo tiene encabezados (Filas < 2)."

        # Iteramos desde la fila 2 (asumiendo que la 1 es encabezado) hasta la última (inclusive).
        for fila in range(2, max_row + 1):
            cell_ref = f"{col}{fila}" # Ej: "A2"
            cell_val = ws[cell_ref].value # Valor actual de la celda
            
            # Convertimos a string para poder analizar el texto, manejando celdas vacías (None)
            val_str = str(cell_val) if cell_val is not None else ""
            
            nuevo_valor_str = val_str
            final_val = None

            # --- LÓGICA SEGÚN EL MODO ---
            
            if mode == "smart_clean":
                # MODO INTELIGENTE (DEFAULT)
                # Mantiene letras, números y espacios. Elimina símbolos raros como @, #, $, etc.
                # También intenta convertir a número real si parece uno.
                
                # Filtro: dejamos pasar solo caracteres alfanuméricos o espacios.
                # 'isalnum()' es True para letras y números.
                nuevo_valor_str = ''.join(filter(lambda x: x.isalnum() or x.isspace(), val_str)).strip()
                
                # Intentamos convertir a Entero si lo que queda parece un número (ignorando espacios internos)
                if nuevo_valor_str.replace(" ", "").isdigit():
                    try: 
                        final_val = int(nuevo_valor_str.replace(" ", ""))
                    except:
                        final_val = nuevo_valor_str # Si falla conversión, dejamos texto
                else:
                    # Si no es número, dejamos el texto limpio.
                    final_val = nuevo_valor_str if nuevo_valor_str else None

            elif mode == "keep_digits":
                # MODO SOLO DÍGITOS ("Deja solo numeros")
                # Filtra solo los dígitos (0-9). Elimina letras y todo lo demás.
                nuevo_valor_str = ''.join(filter(str.isdigit, val_str))
                if nuevo_valor_str:
                    try:
                        final_val = int(nuevo_valor_str) # Convertimos a int
                    except ValueError:
                         final_val = nuevo_valor_str
                else:
                    final_val = None # Si no quedaron dígitos, celda vacía
            
            elif mode == "keep_text":
                # MODO SOLO TEXTO ("Borra numeros")
                # Mantiene todo lo que NO sea dígito.
                nuevo_valor_str = ''.join(filter(lambda x: not x.isdigit(), val_str)).strip()
                final_val = nuevo_valor_str if nuevo_valor_str else None
                
            elif mode == "empty":
                # MODO BORRADO TOTAL ("Borra columna")
                final_val = None

            # --- APLICACIÓN DE CAMBIOS ---
            
            # Comparamos si vale la pena escribir en el Excel (si algo cambió)
            current_val_check = cell_val
            changed = False
            
            # Si tenemos un nuevo valor propuesto...
            if final_val is not None:
                # Si es diferente al anterior...
                if final_val != current_val_check:
                    changed = True
                
                # REGLA EXTRA: Si es un número, forzamos el formato de celda '0' (Número).
                # Esto arregla el problema de "Número almacenado como texto".
                if isinstance(final_val, int):
                     # Si el formato actual es '@' (Texto), lo cambiamos.
                     if ws[cell_ref].number_format == '@':
                         ws[cell_ref].number_format = '0'
                         changed = True
            else:
                # Si el nuevo valor es None (vacío) y antes había algo...
                if current_val_check is not None and current_val_check != "":
                    changed = True

            # Si detectamos cambio, escribimos en la celda
            if changed:
                 ws[cell_ref] = final_val
                 # Aseguramos formato numérico si corresponde
                 if isinstance(final_val, int):
                      ws[cell_ref].number_format = '0'
                 modificadas += 1
            count += 1
            
        mensaje = f"Limpieza ({mode}) de columna {col} completada. Se procesaron {count} filas y se modificaron {modificadas} celdas."
            
    # --- CASO 2: UNIÓN DE COLUMNAS ---
    elif instruccion["action"] == "merge_columns":
        sources = instruccion["sources"] # Lista de columas origen ['A', 'B']
        target = instruccion.get("target")

        # Si no hay destino, calculamos la primera columna libre a la derecha.
        if not target:
            target_idx = ws.max_column + 1
            target = get_column_letter(target_idx)

        # --- Creación de Encabezado Dinámico ---
        # Leemos los títulos de la fila 1 de las columnas origen.
        headers = []
        for col in sources:
            val = ws[f"{col}1"].value
            headers.append(str(val) if val else f"Col{col}")
        
        # Creamos el nuevo título uniendo los originales con " - "
        header_text = f"{' - '.join(headers)}"
        ws[f"{target}1"] = header_text # Escribimos en fila 1 destino

        count = 0
        # Recorremos filas para concatenar valores
        for fila in range(2, ws.max_row + 1):
            valores = []
            for col in sources:
                val = ws[f"{col}{fila}"].value
                valores.append(str(val) if val is not None else "")
            
            # Unimos los valores con espacio
            resultado = " ".join(valores).strip()
            ws[f"{target}{fila}"] = resultado
            count += 1
        
        mensaje = f"Unión de columnas {', '.join(sources)} en columna {target} completada. Header: '{header_text}'. Se generaron {count} filas."

    # Guardamos los cambios en el archivo físico.
    wb.save(file_path)
    wb.close() # Importante: liberar el archivo para que no quede bloqueado
    
    # Devolvemos el mensaje final indicando dónde se guardó
    return f"{mensaje}\nArchivo guardado en: {file_path}"
